"""
app.py - Flask web server for the SLD viewer.

Two pages:
  /                -> the Projects page: a grid of projects you open.
  /project/<pid>   -> the workspace: the project's .xlsx files as tabs, with
                      the single-line diagram for the selected file.

Model: a **project** is a container (a job/site). It holds many **files**
(each uploaded .xlsx is one diagram). Admins create projects, upload files
into them, and rename/delete both. Any logged-in user can open projects,
switch files, edit feeder statuses and export.

Collaboration: each file has a `revision` bumped on every status change;
browsers poll /state and apply other users' edits live. Everything is
persisted (SQLite locally / PostgreSQL on Render) and audited.
"""

import io
import os
import json
import time
import secrets
from datetime import timedelta, datetime
from functools import wraps

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

from flask import (Flask, render_template, jsonify, request, abort, send_file,
                   session, redirect, url_for)
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl import load_workbook

import graph as graph_module
import db

try:
    from PIL import Image, ImageOps
    _PIL = True
except ImportError:                # photos still work, just stored as-is
    _PIL = False

SHEET_NAME = "Energization"
MAX_PHOTOS_PER_FEEDER = 6
PHOTO_MAX_DIM = 1400               # downscale longest side to this many px

# Commissioning workflow stages, in site order. Each feeder gets this
# checklist; every tick records who and when. Admins can add more per project
# (see effective_milestones); these six are always present.
DEFAULT_MILESTONES = ["Cable Pulled", "Glanded", "Terminated",
                      "Tested", "Energized", "Commissioned"]

# Test record types offered in the Tests tab (free 'Other' allowed too).
TEST_TYPES = ["Insulation Resistance (Megger)", "Continuity",
              "Earth / Ground", "Phase Rotation", "Other"]

app = Flask(__name__)

# Render (and most hosts) terminate TLS at a proxy and forward to us over HTTP
# with X-Forwarded-* headers. ProxyFix makes Flask trust those so it knows the
# request is really HTTPS — needed for correct redirects and Secure cookies.
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1)

db.init_db()
app.secret_key = db.get_secret_key()

app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("SECURE_COOKIES", "0") == "1",
    # Sign users out after this much inactivity (the cookie expiry slides
    # forward on each request, so it's effectively an idle timeout).
    PERMANENT_SESSION_LIFETIME=timedelta(hours=int(os.environ.get("SESSION_HOURS", "8"))),
    # Cap request size (workbooks + photos). 25 MB is plenty for either.
    MAX_CONTENT_LENGTH=int(os.environ.get("MAX_UPLOAD_MB", "25")) * 1024 * 1024,
)


@app.context_processor
def inject_asset_helper():
    """`asset('app.js')` -> a static URL with a ?v=<mtime> cache-buster, so
    browsers always pick up new JS/CSS after a deploy."""
    def asset(filename):
        try:
            v = int(os.path.getmtime(os.path.join(app.static_folder, filename)))
        except OSError:
            v = 0
        return url_for("static", filename=filename) + f"?v={v}"
    return {"asset": asset}


def _process_image(raw):
    """Downscale/normalise an uploaded image to keep DB storage small.
    Returns (bytes, mimetype). Falls back to the original if PIL is absent
    or the file can't be parsed as an image."""
    if not _PIL:
        return raw, "application/octet-stream"
    try:
        img = Image.open(io.BytesIO(raw))
        img = ImageOps.exif_transpose(img)        # honour phone orientation
        img = img.convert("RGB")
        img.thumbnail((PHOTO_MAX_DIM, PHOTO_MAX_DIM))
        out = io.BytesIO()
        img.save(out, format="JPEG", quality=80, optimize=True)
        return out.getvalue(), "image/jpeg"
    except Exception:                              # noqa: BLE001 - not an image
        return None, None


# ---------------- CSRF protection ----------------
# Synchronizer-token pattern: a per-session token is rendered into each page
# and must be echoed back in the X-CSRFToken header on every state-changing
# API call. Combined with SameSite=Lax cookies this blocks cross-site writes.
def get_csrf():
    if "csrf" not in session:
        session["csrf"] = secrets.token_hex(16)
    return session["csrf"]


@app.before_request
def _csrf_protect():
    if request.method not in ("POST", "PUT", "PATCH", "DELETE"):
        return
    if request.path == "/login":      # pre-auth form; exempt (login-CSRF is low risk)
        return
    sent = request.headers.get("X-CSRFToken") or request.form.get("csrf_token")
    if not session.get("csrf") or sent != session.get("csrf"):
        return jsonify({"error": "Bad or missing CSRF token. Reload the page and retry."}), 400


# ---------------- Login rate limiting ----------------
# Simple in-memory limiter (fine for our single gunicorn worker): lock out an
# IP after too many failed logins within a window.
LOGIN_FAILS = {}
LOGIN_MAX = int(os.environ.get("LOGIN_MAX_ATTEMPTS", "8"))
LOGIN_WINDOW = 900   # 15 minutes


def _login_blocked(ip):
    now = time.time()
    fails = [t for t in LOGIN_FAILS.get(ip, []) if now - t < LOGIN_WINDOW]
    LOGIN_FAILS[ip] = fails
    return len(fails) >= LOGIN_MAX


def _record_login_fail(ip):
    LOGIN_FAILS.setdefault(ip, []).append(time.time())


@app.route("/healthz")
def healthz():
    """Unauthenticated health check for the hosting platform."""
    return jsonify({"status": "ok"})


# Lazily-built graph per file: fid -> {"data": graphdict, "raw": bytes, "name": str}
FILES = {}


def _build_from_bytes(raw):
    return graph_module.load_and_build(io.BytesIO(raw), SHEET_NAME)


def _apply_overrides(fid, data):
    overrides = db.get_status_overrides(fid)
    if not overrides:
        return
    df = data["_df"]
    for e in data["cytoscape"]["edges"]:
        sn = e["data"]["sn"]
        if sn in overrides:
            st = overrides[sn]
            e["data"]["status"] = st
            e["data"]["color"] = graph_module.STATUS_COLORS.get(st, "#999999")
            df.loc[df["SN"] == sn, "Site status"] = st


def _summary_from_data(data):
    """Count feeders by status for the progress dashboard."""
    counts = {s: 0 for s in graph_module.VALID_STATUSES}
    total = 0
    for e in data["cytoscape"]["edges"]:
        st = e["data"].get("status")
        if st in counts:
            counts[st] += 1
        total += 1
    counts["total"] = total
    return counts


def _save_summary(fid, data):
    """Recompute and persist this file's status counts; returns the dict."""
    summ = _summary_from_data(data)
    db.set_file_summary(fid, json.dumps(summ))
    if fid in FILES:
        FILES[fid]["summary"] = summ
    return summ


def _file(fid):
    """Cached {data, raw, name} for a file, built on first use. None if missing."""
    if fid in FILES:
        return FILES[fid]
    row = db.get_file(fid)
    if not row:
        return None
    raw = bytes(row["data"])
    try:
        data = _build_from_bytes(raw)
    except Exception as e:  # noqa: BLE001
        print(f"Could not build file {fid}: {e}")
        return None
    _apply_overrides(fid, data)
    FILES[fid] = {"data": data, "raw": raw, "name": row["name"], "project_id": row["project_id"]}
    # Backfill the status summary for files uploaded before the dashboard existed.
    if not row.get("summary"):
        _save_summary(fid, data)
    else:
        try:
            FILES[fid]["summary"] = json.loads(row["summary"])
        except (ValueError, TypeError):
            _save_summary(fid, data)
    return FILES[fid]


def _edge_by_sn(data, sn):
    for e in data["cytoscape"]["edges"]:
        if e["data"]["sn"] == sn:
            return e
    return None


# ---------------- Auth ----------------
def current_user():
    return session.get("user")


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            if request.path.startswith("/api/"):
                return jsonify({"error": "Not logged in"}), 401
            return redirect(url_for("login", next=request.path))
        return fn(*args, **kwargs)
    return wrapper


def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u:
            return jsonify({"error": "Not logged in"}), 401
        if u.get("role") != "admin":
            return jsonify({"error": "Admin only"}), 403
        return fn(*args, **kwargs)
    return wrapper


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "GET":
        return render_template("login.html")

    ip = request.remote_addr or "?"
    if _login_blocked(ip):
        return render_template("login.html",
            error="Too many failed attempts. Please wait a few minutes and try again."), 429

    body = request.form if request.form else (request.get_json(silent=True) or {})
    u = db.verify_login((body.get("username") or "").strip(), body.get("password") or "")
    if not u:
        _record_login_fail(ip)
        return render_template("login.html", error="Invalid username or password."), 401

    LOGIN_FAILS.pop(ip, None)
    session.permanent = True               # enable the idle-timeout lifetime
    session["user"] = {"username": u["username"], "role": u["role"]}
    get_csrf()                             # issue a CSRF token for this session
    return redirect(request.args.get("next") or url_for("projects_page"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------- Pages ----------------
@app.route("/")
@login_required
def projects_page():
    default_admin = (current_user()["username"] == db.DEFAULT_ADMIN and
                     db.verify_login(db.DEFAULT_ADMIN, "admin123") is not None)
    return render_template("projects.html", user=current_user(),
                           default_admin_warning=default_admin, csrf=get_csrf())


@app.route("/project/<int:pid>")
@login_required
def project_page(pid):
    proj = db.get_project(pid)
    if not proj:
        return redirect(url_for("projects_page"))
    return render_template("project.html", user=current_user(), project=dict(proj), csrf=get_csrf())


@app.route("/api/account/password", methods=["POST"])
@login_required
def api_change_password():
    """Let a logged-in user change their own password."""
    body = request.get_json(silent=True) or {}
    me = current_user()["username"]
    if not db.verify_login(me, body.get("current_password") or ""):
        return jsonify({"error": "Current password is incorrect."}), 400
    new = body.get("new_password") or ""
    if len(new) < 6:
        return jsonify({"error": "New password must be at least 6 characters."}), 400
    db.set_password(me, new)
    db.log("password_change_self", me)
    return jsonify({"ok": True})


@app.route("/api/me")
@login_required
def api_me():
    return jsonify(current_user())


# ---------------- Projects API ----------------
@app.route("/api/projects", methods=["GET"])
@login_required
def api_projects():
    # Aggregate each project's feeder status counts across its files, so the
    # cards can show a progress bar without parsing any workbooks.
    agg = {}
    for r in db.all_file_summaries():
        if not r["summary"]:
            continue
        try:
            s = json.loads(r["summary"])
        except (ValueError, TypeError):
            continue
        a = agg.setdefault(r["project_id"], {})
        for k, v in s.items():
            a[k] = a.get(k, 0) + v
    projects = []
    for r in db.list_projects():
        d = dict(r)
        d["progress"] = agg.get(r["id"], {})
        projects.append(d)
    return jsonify({"projects": projects, "can_manage": current_user()["role"] == "admin"})


@app.route("/api/projects", methods=["POST"])
@admin_required
def api_create_project():
    body = request.get_json(silent=True) or {}
    try:
        pid = db.create_project(body.get("name", ""), current_user()["username"])
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    db.log("project_create", current_user()["username"], project_id=pid,
           new_value=body.get("name"))
    return jsonify({"ok": True, "id": pid})


@app.route("/api/projects/<int:pid>", methods=["PATCH"])
@admin_required
def api_rename_project(pid):
    if not db.project_exists(pid):
        return jsonify({"error": "No such project"}), 404
    body = request.get_json(silent=True) or {}
    try:
        db.rename_project(pid, body.get("name", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    db.log("project_rename", current_user()["username"], project_id=pid, new_value=body.get("name"))
    return jsonify({"ok": True})


@app.route("/api/projects/<int:pid>", methods=["DELETE"])
@admin_required
def api_delete_project(pid):
    if not db.project_exists(pid):
        return jsonify({"error": "No such project"}), 404
    name = db.get_project(pid)["name"]
    # Drop any cached files belonging to this project.
    for fid in [r["id"] for r in db.list_files(pid)]:
        FILES.pop(fid, None)
    db.delete_project(pid)
    db.log("project_delete", current_user()["username"], paulos=name)
    return jsonify({"ok": True})


# ---------------- Files API ----------------
@app.route("/api/projects/<int:pid>/files", methods=["GET"])
@login_required
def api_list_files(pid):
    proj = db.get_project(pid)
    if not proj:
        return jsonify({"error": "No such project"}), 404
    return jsonify({
        "project": dict(proj),
        "files": [dict(r) for r in db.list_files(pid)],
        "can_manage": current_user()["role"] == "admin",
    })


@app.route("/api/projects/<int:pid>/files", methods=["POST"])
@admin_required
def api_upload_file(pid):
    if not db.project_exists(pid):
        return jsonify({"error": "No such project"}), 404
    f = request.files.get("file")
    if f is None or f.filename == "":
        return jsonify({"error": "No file uploaded."}), 400
    if not f.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Please upload an .xlsx file."}), 400
    raw = f.read()
    try:
        data = _build_from_bytes(raw)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:  # noqa: BLE001
        return jsonify({"error": f"Could not read that file: {e}"}), 400

    name = (request.form.get("name") or "").strip() or f.filename.rsplit(".", 1)[0]
    who = current_user()["username"]
    fid = db.create_file(pid, name, f.filename, raw, who)
    FILES[fid] = {"data": data, "raw": raw, "name": name, "project_id": pid}
    _save_summary(fid, data)
    db.log("file_upload", who, project_id=pid, file_id=fid, paulos=f.filename,
           new_value=f"{data['stats']['panels']} panels, {data['stats']['feeders']} feeders")
    return jsonify({"ok": True, "id": fid, "name": name, "stats": data["stats"]})


@app.route("/api/files/<int:fid>", methods=["PATCH"])
@admin_required
def api_rename_file(fid):
    meta = db.get_file_meta(fid)
    if not meta:
        return jsonify({"error": "No such file"}), 404
    body = request.get_json(silent=True) or {}
    try:
        db.rename_file(fid, body.get("name", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if fid in FILES:
        FILES[fid]["name"] = body["name"].strip()
    db.log("file_rename", current_user()["username"], project_id=meta["project_id"],
           file_id=fid, new_value=body.get("name"))
    return jsonify({"ok": True})


@app.route("/api/files/<int:fid>", methods=["DELETE"])
@admin_required
def api_delete_file(fid):
    meta = db.get_file_meta(fid)
    if not meta:
        return jsonify({"error": "No such file"}), 404
    db.delete_file(fid)
    FILES.pop(fid, None)
    db.log("file_delete", current_user()["username"], project_id=meta["project_id"],
           paulos=meta["name"])
    return jsonify({"ok": True})


# ---------------- Per-file graph / data ----------------
@app.route("/api/files/<int:fid>/graph")
@login_required
def api_graph(fid):
    proj = _file(fid)
    if proj is None:
        return jsonify({"loaded": False}), 404
    data = proj["data"]
    return jsonify({
        "loaded": True, "name": proj["name"], "revision": db.get_revision(fid),
        "elements": data["cytoscape"], "findings": data["findings"], "stats": data["stats"],
        "statuses": graph_module.VALID_STATUSES, "colors": graph_module.STATUS_COLORS,
    })


@app.route("/api/files/<int:fid>/node/<path:node_id>")
@login_required
def api_node(fid, node_id):
    proj = _file(fid)
    if proj is None:
        abort(404)
    data = proj["data"]
    incoming, outgoing = [], []
    for e in data["cytoscape"]["edges"]:
        d = e["data"]
        if d["target"] == node_id:
            incoming.append(d)
        if d["source"] == node_id:
            outgoing.append(d)
    if not incoming and not outgoing:
        if not any(n["data"]["id"] == node_id for n in data["cytoscape"]["nodes"]):
            abort(404)
    return jsonify({"id": node_id, "incoming": incoming, "outgoing": outgoing})


@app.route("/api/files/<int:fid>/edge/<int:sn>/status", methods=["POST"])
@login_required
def api_set_status(fid, sn):
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    body = request.get_json(silent=True) or {}
    new_status = body.get("status")
    if new_status not in graph_module.VALID_STATUSES:
        return jsonify({"error": f"Invalid status. Must be one of: {graph_module.VALID_STATUSES}"}), 400
    data = proj["data"]
    edge = _edge_by_sn(data, sn)
    if edge is None:
        return jsonify({"error": f"No feeder with SN {sn}"}), 404

    old_status = edge["data"]["status"]
    edge["data"]["status"] = new_status
    edge["data"]["color"] = graph_module.STATUS_COLORS[new_status]
    data["_df"].loc[data["_df"]["SN"] == sn, "Site status"] = new_status

    who = current_user()["username"]
    db.set_feeder_status(fid, sn, new_status, who)
    if old_status != new_status:
        db.log("status_change", who, project_id=proj["project_id"], file_id=fid, sn=sn,
               paulos=edge["data"].get("paulos"), source=edge["data"].get("source"),
               target=edge["data"].get("target"), old_value=old_status, new_value=new_status)
    _save_summary(fid, data)   # keep the dashboard counts current
    return jsonify({"ok": True, "sn": sn, "status": new_status,
                    "color": graph_module.STATUS_COLORS[new_status], "revision": db.get_revision(fid)})


@app.route("/api/files/<int:fid>/bulk-status", methods=["POST"])
@login_required
def api_bulk_status(fid):
    """Set the same status on many feeders at once (e.g. energize a whole board)."""
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    body = request.get_json(silent=True) or {}
    new_status = body.get("status")
    sns = body.get("sns") or []
    if new_status not in graph_module.VALID_STATUSES:
        return jsonify({"error": f"Invalid status. Must be one of: {graph_module.VALID_STATUSES}"}), 400

    data = proj["data"]
    who = current_user()["username"]
    changed = 0
    for sn in sns:
        edge = _edge_by_sn(data, sn)
        if edge is None:
            continue
        old_status = edge["data"]["status"]
        edge["data"]["status"] = new_status
        edge["data"]["color"] = graph_module.STATUS_COLORS[new_status]
        data["_df"].loc[data["_df"]["SN"] == sn, "Site status"] = new_status
        db.set_feeder_status(fid, sn, new_status, who)
        if old_status != new_status:
            changed += 1
            db.log("status_change", who, project_id=proj["project_id"], file_id=fid, sn=sn,
                   paulos=edge["data"].get("paulos"), source=edge["data"].get("source"),
                   target=edge["data"].get("target"), old_value=old_status, new_value=new_status)
    _save_summary(fid, data)
    return jsonify({"ok": True, "changed": changed, "status": new_status,
                    "color": graph_module.STATUS_COLORS[new_status], "revision": db.get_revision(fid)})


@app.route("/api/files/<int:fid>/notes")
@login_required
def api_notes(fid):
    if not db.get_file_meta(fid):
        return jsonify({"error": "No such file"}), 404
    return jsonify({"notes": db.get_notes(fid)})


@app.route("/api/files/<int:fid>/edge/<int:sn>/note", methods=["POST"])
@login_required
def api_set_note(fid, sn):
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    edge = _edge_by_sn(proj["data"], sn)
    if edge is None:
        return jsonify({"error": f"No feeder with SN {sn}"}), 404
    note = (request.get_json(silent=True) or {}).get("note", "")
    who = current_user()["username"]
    db.set_note(fid, sn, note, who)
    db.log("note", who, project_id=proj["project_id"], file_id=fid, sn=sn,
           paulos=edge["data"].get("paulos"), new_value=(note or "")[:120])
    return jsonify({"ok": True, "sn": sn, "note": (note or "").strip(), "by": who})


# ---------------- Feeder photos ----------------
@app.route("/api/files/<int:fid>/photos")
@login_required
def api_list_photos(fid):
    if not db.get_file_meta(fid):
        return jsonify({"error": "No such file"}), 404
    return jsonify({"photos": db.list_photos(fid)})


@app.route("/api/files/<int:fid>/edge/<int:sn>/photos", methods=["POST"])
@login_required
def api_add_photo(fid, sn):
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    edge = _edge_by_sn(proj["data"], sn)
    if edge is None:
        return jsonify({"error": f"No feeder with SN {sn}"}), 404
    if db.count_photos(fid, sn) >= MAX_PHOTOS_PER_FEEDER:
        return jsonify({"error": f"Limit of {MAX_PHOTOS_PER_FEEDER} photos per feeder reached."}), 400
    f = request.files.get("photo")
    if f is None or f.filename == "":
        return jsonify({"error": "No photo uploaded."}), 400

    data, mime = _process_image(f.read())
    if data is None:
        return jsonify({"error": "That doesn't look like an image."}), 400

    who = current_user()["username"]
    caption = request.form.get("caption", "")
    pid_ = db.add_photo(fid, sn, mime, data, caption, who)
    db.log("photo_add", who, project_id=proj["project_id"], file_id=fid, sn=sn,
           paulos=edge["data"].get("paulos"), new_value=(caption or "")[:120])
    return jsonify({"ok": True, "id": pid_, "sn": sn, "caption": (caption or "").strip(), "by": who})


@app.route("/api/photos/<int:photo_id>")
@login_required
def api_get_photo(photo_id):
    row = db.get_photo(photo_id)
    if not row:
        abort(404)
    resp = send_file(io.BytesIO(bytes(row["data"])), mimetype=row["mimetype"],
                     download_name=f"feeder-{row['sn']}-{photo_id}.jpg")
    resp.headers["Cache-Control"] = "private, max-age=86400"
    return resp


@app.route("/api/photos/<int:photo_id>", methods=["DELETE"])
@login_required
def api_delete_photo(photo_id):
    meta = db.get_photo_meta(photo_id)
    if not meta:
        return jsonify({"error": "No such photo"}), 404
    # Only the uploader or an admin may delete a photo.
    me = current_user()
    if me["role"] != "admin" and meta["uploaded_by"] != me["username"]:
        return jsonify({"error": "Only the uploader or an admin can delete this photo."}), 403
    db.delete_photo(photo_id)
    db.log("photo_delete", me["username"], file_id=meta["file_id"], sn=meta["sn"])
    return jsonify({"ok": True})


# ---------------- Commissioning workflow (milestones) ----------------
def effective_milestones(pid):
    """The default stages plus any the admin added to this project."""
    return DEFAULT_MILESTONES + [r["name"] for r in db.get_project_milestones(pid)]


@app.route("/api/files/<int:fid>/workflow")
@login_required
def api_workflow(fid):
    meta = db.get_file_meta(fid)
    if not meta:
        return jsonify({"error": "No such file"}), 404
    custom = [dict(r) for r in db.get_project_milestones(meta["project_id"])]
    return jsonify({
        "milestones": DEFAULT_MILESTONES + [c["name"] for c in custom],
        "custom": custom,                         # [{id, name}] removable by admins
        "done": db.get_milestones(fid),
        "can_manage": current_user()["role"] == "admin",
    })


@app.route("/api/projects/<int:pid>/milestones", methods=["POST"])
@admin_required
def api_add_project_milestone(pid):
    if not db.project_exists(pid):
        return jsonify({"error": "No such project"}), 404
    try:
        mid = db.add_project_milestone(pid, (request.get_json(silent=True) or {}).get("name", ""))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    db.log("milestone_add", current_user()["username"], project_id=pid,
           new_value=(request.get_json(silent=True) or {}).get("name"))
    return jsonify({"ok": True, "id": mid})


@app.route("/api/milestones/<int:mid>", methods=["DELETE"])
@admin_required
def api_delete_project_milestone(mid):
    meta = db.get_milestone_meta(mid)
    if not meta:
        return jsonify({"error": "No such milestone"}), 404
    db.delete_project_milestone(mid)
    db.log("milestone_delete", current_user()["username"], project_id=meta["project_id"],
           new_value=meta["name"])
    return jsonify({"ok": True})


@app.route("/api/files/<int:fid>/edge/<int:sn>/milestone", methods=["POST"])
@login_required
def api_set_milestone(fid, sn):
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    edge = _edge_by_sn(proj["data"], sn)
    if edge is None:
        return jsonify({"error": f"No feeder with SN {sn}"}), 404
    body = request.get_json(silent=True) or {}
    milestone = body.get("milestone")
    done = bool(body.get("done"))
    valid = effective_milestones(proj["project_id"])
    if milestone not in valid:
        return jsonify({"error": f"Invalid milestone. Must be one of: {valid}"}), 400

    who = current_user()["username"]
    db.set_milestone(fid, sn, milestone, done, who)
    db.log("milestone", who, project_id=proj["project_id"], file_id=fid, sn=sn,
           paulos=edge["data"].get("paulos"),
           old_value=("" if done else milestone), new_value=(milestone if done else ""))
    return jsonify({"ok": True, "sn": sn, "milestone": milestone, "done": done, "by": who})


# ---------------- Test records ----------------
@app.route("/api/files/<int:fid>/tests")
@login_required
def api_tests(fid):
    if not db.get_file_meta(fid):
        return jsonify({"error": "No such file"}), 404
    return jsonify({"types": TEST_TYPES, "tests": db.get_tests(fid)})


@app.route("/api/files/<int:fid>/edge/<int:sn>/tests", methods=["POST"])
@login_required
def api_add_test(fid, sn):
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    edge = _edge_by_sn(proj["data"], sn)
    if edge is None:
        return jsonify({"error": f"No feeder with SN {sn}"}), 404
    body = request.get_json(silent=True) or {}
    test_type = (body.get("test_type") or "").strip()
    result = body.get("result")
    if not test_type:
        return jsonify({"error": "Test type is required."}), 400
    if result not in ("Pass", "Fail"):
        return jsonify({"error": "Result must be Pass or Fail."}), 400

    who = current_user()["username"]
    tid = db.add_test(fid, sn, test_type, body.get("value"), result, body.get("notes"), who)
    db.log("test", who, project_id=proj["project_id"], file_id=fid, sn=sn,
           paulos=edge["data"].get("paulos"), source=test_type,
           new_value=f"{result}" + (f" ({body.get('value')})" if body.get("value") else ""))
    return jsonify({"ok": True, "id": tid, "sn": sn})


@app.route("/api/tests/<int:test_id>", methods=["DELETE"])
@login_required
def api_delete_test(test_id):
    meta = db.get_test_meta(test_id)
    if not meta:
        return jsonify({"error": "No such test"}), 404
    me = current_user()
    if me["role"] != "admin" and meta["tested_by"] != me["username"]:
        return jsonify({"error": "Only the tester or an admin can delete this record."}), 403
    db.delete_test(test_id)
    db.log("test_delete", me["username"], file_id=meta["file_id"], sn=meta["sn"])
    return jsonify({"ok": True})


# ---------------- Punch list / snags ----------------
@app.route("/api/files/<int:fid>/snags")
@login_required
def api_snags(fid):
    if not db.get_file_meta(fid):
        return jsonify({"error": "No such file"}), 404
    return jsonify({"snags": [dict(r) for r in db.get_snags(fid)]})


@app.route("/api/files/<int:fid>/edge/<int:sn>/snags", methods=["POST"])
@login_required
def api_add_snag(fid, sn):
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    edge = _edge_by_sn(proj["data"], sn)
    if edge is None:
        return jsonify({"error": f"No feeder with SN {sn}"}), 404
    body = request.get_json(silent=True) or {}
    who = current_user()["username"]
    try:
        sid = db.add_snag(fid, sn, body.get("description"), who)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    db.log("snag_open", who, project_id=proj["project_id"], file_id=fid, sn=sn,
           paulos=edge["data"].get("paulos"), new_value=(body.get("description") or "")[:120])
    return jsonify({"ok": True, "id": sid, "sn": sn})


@app.route("/api/snags/<int:snag_id>", methods=["PATCH"])
@login_required
def api_update_snag(snag_id):
    snag = db.get_snag(snag_id)
    if not snag:
        return jsonify({"error": "No such snag"}), 404
    status = (request.get_json(silent=True) or {}).get("status")
    if status not in ("open", "closed"):
        return jsonify({"error": "Status must be open or closed."}), 400
    who = current_user()["username"]
    db.set_snag_status(snag_id, status, who)
    db.log("snag_" + ("close" if status == "closed" else "reopen"), who,
           file_id=snag["file_id"], sn=snag["sn"],
           new_value=snag["description"][:120])
    return jsonify({"ok": True, "status": status})


# ---------------- Printable progress report ----------------
@app.route("/project/<int:pid>/report")
@login_required
def project_report(pid):
    proj = db.get_project(pid)
    if not proj:
        return redirect(url_for("projects_page"))

    files_data = []
    total = {s: 0 for s in graph_module.VALID_STATUSES}
    total["feeders"] = 0
    open_snags = []
    for f in db.list_files(pid):
        fid = f["id"]
        summ = json.loads(f["summary"]) if f["summary"] else {}
        boards = []
        cached = _file(fid)
        if cached:
            per_board = {}
            for e in cached["data"]["cytoscape"]["edges"]:
                d = e["data"]
                b = per_board.setdefault(d["source"], {s: 0 for s in graph_module.VALID_STATUSES})
                b[d["status"]] = b.get(d["status"], 0) + 1
            boards = [{"name": k, **v, "total": sum(v.values())}
                      for k, v in sorted(per_board.items())]
        for s in graph_module.VALID_STATUSES:
            total[s] += summ.get(s, 0)
        total["feeders"] += summ.get("total", 0)
        for sn in db.get_snags(fid):
            if sn["status"] == "open":
                open_snags.append({**dict(sn), "file": f["name"]})
        files_data.append({
            "name": f["name"], "summary": summ, "boards": boards,
            "milestones": db.milestone_counts(fid),
            "tests": db.test_counts(fid),
            "snags": db.snag_counts(fid),
        })

    pct = round(total["Energized"] / total["feeders"] * 100) if total["feeders"] else 0
    db.log("report", current_user()["username"], project_id=pid)
    return render_template("report.html", project=dict(proj), files=files_data,
                           total=total, pct=pct, open_snags=open_snags,
                           milestones=effective_milestones(pid), colors=graph_module.STATUS_COLORS,
                           statuses=graph_module.VALID_STATUSES,
                           user=current_user(), now=datetime.now())


@app.route("/api/files/<int:fid>/state")
@login_required
def api_state(fid):
    rev = db.get_revision(fid)
    if rev is None:
        return jsonify({"error": "No such file"}), 404
    return jsonify({"revision": rev, "statuses": db.get_status_overrides(fid)})


# ---------------- Presence ("who's viewing") ----------------
# In-memory registry of who is currently looking at each file. Accurate with
# our single worker; each client refreshes it on every poll tick.
PRESENCE = {}
PRESENCE_TTL = 20      # seconds a viewer stays "present" between pings


@app.route("/api/files/<int:fid>/presence")
@login_required
def api_presence(fid):
    now = time.time()
    me = current_user()["username"]
    book = PRESENCE.setdefault(fid, {})
    book[me] = now
    for u in [u for u, t in book.items() if now - t > PRESENCE_TTL]:
        del book[u]
    if not book:
        PRESENCE.pop(fid, None)
    return jsonify({"viewers": sorted(book.keys()), "you": me})


@app.route("/api/files/<int:fid>/history")
@login_required
def api_history(fid):
    return jsonify({"entries": [dict(r) for r in db.recent_audit(file_id=fid)]})


@app.route("/api/files/<int:fid>/export")
@login_required
def api_export(fid):
    proj = _file(fid)
    if proj is None:
        return jsonify({"error": "No such file"}), 404
    data, raw = proj["data"], proj["raw"]
    df = data["_df"]
    status_by_sn = {int(sn): st for sn, st in zip(df["SN"], df["Site status"])}
    wb = load_workbook(io.BytesIO(raw))
    ws = wb[SHEET_NAME]
    hdr = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    sn_col, st_col = hdr["SN"], hdr["Site status"]
    for r in range(2, ws.max_row + 1):
        sn = ws.cell(row=r, column=sn_col).value
        if sn is not None and int(sn) in status_by_sn:
            ws.cell(row=r, column=st_col).value = status_by_sn[int(sn)]
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    db.log("export", current_user()["username"], project_id=proj["project_id"], file_id=fid)
    safe = (proj["name"] or "file").replace(" ", "_")
    return send_file(buf, as_attachment=True, download_name=f"{safe}_UPDATED.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ---------------- User management (admin only) ----------------
@app.route("/api/users", methods=["GET"])
@admin_required
def api_users():
    return jsonify({"users": [dict(r) for r in db.list_users()]})


@app.route("/api/users", methods=["POST"])
@admin_required
def api_create_user():
    body = request.get_json(silent=True) or {}
    try:
        db.create_user(body.get("username", ""), body.get("password", ""), body.get("role", "user"))
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    db.log("user_create", current_user()["username"], paulos=body.get("username"),
           new_value=body.get("role", "user"))
    return jsonify({"ok": True})


@app.route("/api/users/<username>", methods=["PATCH"])
@admin_required
def api_update_user(username):
    body = request.get_json(silent=True) or {}
    me = current_user()["username"]
    if "role" in body:
        target = db.get_user(username)
        if (target and target["role"] == "admin" and body["role"] != "admin"
                and db.count_admins() <= 1):
            return jsonify({"error": "Can't remove the last admin."}), 400
        try:
            db.set_role(username, body["role"])
        except ValueError as e:
            return jsonify({"error": str(e)}), 400
        db.log("user_role", me, paulos=username, new_value=body["role"])
    if body.get("password"):
        db.set_password(username, body["password"])
        db.log("user_password", me, paulos=username)
    return jsonify({"ok": True})


@app.route("/api/users/<username>", methods=["DELETE"])
@admin_required
def api_delete_user(username):
    if username == current_user()["username"]:
        return jsonify({"error": "You can't delete your own account."}), 400
    target = db.get_user(username)
    if target and target["role"] == "admin" and db.count_admins() <= 1:
        return jsonify({"error": "Can't delete the last admin."}), 400
    db.delete_user(username)
    db.log("user_delete", current_user()["username"], paulos=username)
    return jsonify({"ok": True})


if __name__ == "__main__":
    host = os.environ.get("HOST", "127.0.0.1")
    port = int(os.environ.get("PORT", "5000"))
    debug = os.environ.get("FLASK_DEBUG", "1") == "1"
    app.run(host=host, port=port, debug=debug)
